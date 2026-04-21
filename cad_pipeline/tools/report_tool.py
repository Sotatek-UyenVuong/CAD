"""report_tool.py — Generate PDF, DOCX, or Excel reports from chat context.

Flow (PDF):
  context_md pages + query + tool_result
    → Gemini Flash: draft structured Markdown report
    → write .md temp file
    → md_export.convert_to_pdf  (Pandoc + xelatex)
    → return {format, file_path, file_name, markdown}

Flow (DOCX):
  context_md pages + query + tool_result
    → Gemini Flash: draft structured Markdown report
    → write .md temp file
    → md_export.convert_to_docx (Pandoc)
    → return {format, file_path, file_name, markdown}

Flow (Excel):
  context_md pages + query + tool_result
    → Gemini Flash: extract structured JSON (summary, pages, tables)
    → build multi-sheet Excel via openpyxl
    → return {format, file_path, file_name}
"""

from __future__ import annotations

import json
import re
import uuid
from datetime import datetime
from pathlib import Path

from cad_pipeline.config import GEMINI_API_KEY, GEMINI_FLASH_MODEL, REPORTS_DIR


# ── Markdown report builder ─────────────────────────────────────────────────

_REPORT_SYSTEM = """\
You are a professional architectural document analyst.
Write a structured Markdown report in the SAME LANGUAGE as the user query
(Vietnamese, Japanese, or English).

Report structure:
# {TITLE}

## Summary
(2–3 sentences summarising what was found)

## Details
(Use tables or bullet lists. Include page references.)

## Results
(Numerical results, counts, areas, or search hits if available)

## Source Pages
| Page | File | Summary |
|---|---|---|
| ... | ... | ... |

Only include sections that have content. Be concise and accurate.
Do NOT add any preamble — output the Markdown directly.
Important quality rules:
- Use section titles and prose in ONLY one language (the user's language). Do not mix multilingual labels in headings.
- Use clean Markdown bullets with "-" (avoid "*").
- For tables, output valid Markdown table rows only (no broken pipes, no placeholder separator duplicates).
- Do not include raw prompt artifacts like "chat_history" in the final prose; instead explain concrete findings.
"""


def _build_report_markdown(
    query: str,
    pages: list[dict],
    tool_result: dict | None,
    model: str | None = None,
) -> str:
    """Ask Gemini Flash to draft a structured Markdown report."""
    from google import genai  # type: ignore

    _model = model or GEMINI_FLASH_MODEL
    client = genai.Client(api_key=GEMINI_API_KEY)

    pages_text = "\n\n".join(
        f"=== PAGE {p.get('page_number', '?')} ({p.get('file_name', '')}) ===\n"
        f"{p.get('short_summary', '')}\n"
        f"{p.get('context_md', '')[:2000]}"
        for p in pages[:5]
    )

    tool_text = ""
    if tool_result:
        tool_text = f"\n\nTool result:\n{json.dumps(tool_result, ensure_ascii=False, indent=2)[:1500]}"

    user_prompt = (
        f"{_REPORT_SYSTEM}\n\n"
        f"User query: {query}\n\n"
        f"Page content:\n{pages_text}"
        f"{tool_text}"
    )

    try:
        response = client.models.generate_content(model=_model, contents=user_prompt)
        return response.text.strip()
    except Exception as exc:
        # Fallback: minimal report from page summaries
        lines = [f"# Report: {query}", "", "## Source Pages", ""]
        for p in pages:
            lines.append(f"- **Page {p.get('page_number')}** ({p.get('file_name', '')}): "
                         f"{p.get('short_summary', '')[:150]}")
        if tool_result:
            lines += ["", "## Tool Result", "", f"```json", json.dumps(tool_result, ensure_ascii=False, indent=2), "```"]
        lines += ["", f"*Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}*"]
        return "\n".join(lines)


# ── PDF export ──────────────────────────────────────────────────────────────

def run_report_pdf(
    query: str,
    pages: list[dict],
    tool_result: dict | None = None,
    output_dir: str | Path | None = None,
    title: str | None = None,
    model: str | None = None,
) -> dict:
    """Generate a PDF report from context pages and tool result.

    Returns:
        {
          "format": "pdf",
          "file_path": str,
          "file_name": str,
          "markdown": str,      # the generated Markdown source
          "success": bool,
          "error": str | None,
        }
    """
    from cad_pipeline.tools.md_export import convert_to_pdf, inject_front_matter

    markdown = _build_report_markdown(query, pages, tool_result, model=model)

    _title = title or f"CAD Report — {query[:60]}"
    _date  = datetime.now().strftime("%Y-%m-%d")
    _out_dir = Path(output_dir) if output_dir else REPORTS_DIR
    _out_dir.mkdir(parents=True, exist_ok=True)

    slug = re.sub(r"[^\w\-]", "_", query[:40]).strip("_")
    uniq = datetime.now().strftime("%Y%m%d_%H%M%S") + "_" + uuid.uuid4().hex[:6]
    md_path  = _out_dir / f"report_{slug}_{uniq}.md"
    pdf_path = _out_dir / f"report_{slug}_{uniq}.pdf"

    md_path.write_text(markdown, encoding="utf-8")

    success = convert_to_pdf(
        md_path=md_path,
        out_path=pdf_path,
        title=_title,
        author="CAD Pipeline",
        date=_date,
        toc=False,
    )

    if success and pdf_path.exists():
        # Keep folder clean: PDF is the final artifact, markdown stays in response only.
        md_path.unlink(missing_ok=True)
        return {
            "format":    "pdf",
            "file_path": str(pdf_path),
            "file_name": pdf_path.name,
            "markdown":  markdown,
            "success":   True,
            "error":     None,
        }

    # Fallback: return md if PDF failed
    return {
        "format":    "pdf",
        "file_path": str(md_path),
        "file_name": md_path.name,
        "markdown":  markdown,
        "success":   False,
        "error":     "PDF conversion failed; Markdown source available.",
    }


def run_report_docx(
    query: str,
    pages: list[dict],
    tool_result: dict | None = None,
    output_dir: str | Path | None = None,
    title: str | None = None,
    model: str | None = None,
) -> dict:
    """Generate a DOCX report from context pages and tool result."""
    from cad_pipeline.tools.md_export import convert_to_docx

    markdown = _build_report_markdown(query, pages, tool_result, model=model)

    _title = title or f"CAD Report — {query[:60]}"
    _date = datetime.now().strftime("%Y-%m-%d")
    _out_dir = Path(output_dir) if output_dir else REPORTS_DIR
    _out_dir.mkdir(parents=True, exist_ok=True)

    slug = re.sub(r"[^\w\-]", "_", query[:40]).strip("_")
    uniq = datetime.now().strftime("%Y%m%d_%H%M%S") + "_" + uuid.uuid4().hex[:6]
    md_path = _out_dir / f"report_{slug}_{uniq}.md"
    docx_path = _out_dir / f"report_{slug}_{uniq}.docx"

    md_path.write_text(markdown, encoding="utf-8")
    template_dir = Path(__file__).resolve().parent / "templates"
    reference_doc = template_dir / "cad_report_template.docx"
    success = convert_to_docx(
        md_path=md_path,
        out_path=docx_path,
        title=_title,
        author="CAD Pipeline",
        date=_date,
        toc=False,
        reference_doc=reference_doc,
    )

    if success and docx_path.exists():
        md_path.unlink(missing_ok=True)
        return {
            "format": "docx",
            "file_path": str(docx_path),
            "file_name": docx_path.name,
            "markdown": markdown,
            "success": True,
            "error": None,
        }

    return {
        "format": "docx",
        "file_path": str(md_path),
        "file_name": md_path.name,
        "markdown": markdown,
        "success": False,
        "error": "DOCX conversion failed; Markdown source available.",
    }


# ── Excel export ─────────────────────────────────────────────────────────────

def _build_report_json(
    query: str,
    pages: list[dict],
    tool_result: dict | None,
    answer_text: str | None = None,
    chat_history: list[dict] | None = None,
    model: str | None = None,
) -> dict:
    """Ask Gemini Flash to extract dynamic table schema for Excel export.

    Strategy:
    - Prefer chat history when user already provided structured info in conversation.
    - Use page context when chat information is insufficient.
    - Use hybrid mode when both are needed.
    """
    from google import genai  # type: ignore

    _model = model or GEMINI_FLASH_MODEL
    client = genai.Client(api_key=GEMINI_API_KEY)

    pages_text = "\n\n".join(
        (
            f"=== PAGE {p.get('page_number')} | {p.get('file_name', '')} ===\n"
            f"Summary: {p.get('short_summary', '')}\n"
            f"Context:\n{(p.get('context_md', '') or '')[:2500]}"
        )
        for p in pages[:8]
    )

    history_lines: list[str] = []
    if chat_history:
        for turn in chat_history[-8:]:
            history_lines.append(f"User: {turn.get('role_user', '')}")
            history_lines.append(f"Assistant: {turn.get('role_assistant', '')}")
    chat_text = "\n".join(history_lines)

    tool_text = json.dumps(tool_result or {}, ensure_ascii=False, indent=2)[:4000]
    ans_text = (answer_text or "").strip()

    prompt = f"""You are generating an Excel report schema and data for a CAD assistant.
Decide data source strategy first:
- "chat_history": if user/assistant conversation already contains the requested values.
- "page_context": if values must be extracted from pages/context.
- "hybrid": if both are needed.

Extract dynamic Excel-ready table data from this CAD query response.
Query: {query}

Recent chat history:
{chat_text}

Pages (summary + context):
{pages_text}

Current answer text:
{ans_text}

Tool result JSON:
{tool_text}

Return ONLY JSON with this structure:
{{
  "title": "<report title>",
  "source_strategy": "chat_history" | "page_context" | "hybrid",
  "decision_note": "<one sentence on why this strategy>",
  "summary": "<2 sentences>",
  "pages": [
    {{"page_number": <int>, "file_name": "<str>", "summary": "<str>"}},
    ...
  ],
  "tables": [
    {{
      "sheet_name": "<short sheet name>",
      "columns": ["<col1>", "<col2>", "..."],
      "rows": [
        {{"<col1>": "<value>", "<col2>": "<value>"}},
        ...
      ]
    }},
    ...
  ]
}}
Rules:
- Infer report title/sheets/columns from user intent + conversation context.
- If chat history already contains complete values, do NOT fabricate page-derived rows.
- If values are missing in chat, use page context to fill rows.
- Build tables from provided inputs only (no hallucinations).
- Prefer numeric types for numeric values.
- Keep sheet_name <= 25 chars.
- Keep column names concise and stable (no markdown, no multilingual slash labels).
- Return at least one meaningful row; avoid placeholder-only rows unless there is truly no extractable data.
Output raw JSON only."""

    last_err: Exception | None = None
    for _attempt in range(3):
        try:
            retry_note = ""
            if _attempt > 0:
                retry_note = (
                    "\n\nIMPORTANT RETRY CONSTRAINT:\n"
                    "- Previous response was invalid/empty.\n"
                    "- You MUST return at least 1 table in `tables`.\n"
                    "- If there is little structured data, create one sheet named "
                    "`ConversationExtract` with columns [\"field\", \"value\"].\n"
                    "- Always include `source_strategy` and `decision_note`.\n"
                    "- Output JSON only.\n"
                )
            response = client.models.generate_content(model=_model, contents=prompt + retry_note)
            raw = response.text.strip()
            raw = re.sub(r"^```[a-z]*\n?", "", raw).rstrip("`").strip()
            parsed = json.loads(raw)
            tables = parsed.get("tables", [])
            if not isinstance(tables, list) or len(tables) == 0:
                raise ValueError("Gemini returned no tables.")
            if parsed.get("source_strategy") not in {"chat_history", "page_context", "hybrid"}:
                raise ValueError("Gemini returned invalid source_strategy.")
            return parsed
        except Exception as exc:  # retry on any parse/schema/model error
            last_err = exc
            continue

    raise RuntimeError(f"Gemini table generation failed after retries: {last_err}")


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """Excel safe sheet name: <=31 chars, no special chars, unique."""
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", (name or "Data")).strip() or "Data"
    cleaned = cleaned[:31]
    base = cleaned
    idx = 2
    while cleaned in used:
        suffix = f"_{idx}"
        cleaned = f"{base[:31-len(suffix)]}{suffix}"
        idx += 1
    used.add(cleaned)
    return cleaned


def run_report_excel(
    query: str,
    pages: list[dict],
    tool_result: dict | None = None,
    answer_text: str | None = None,
    chat_history: list[dict] | None = None,
    output_dir: str | Path | None = None,
    model: str | None = None,
) -> dict:
    """Generate an Excel report from context pages and tool result.

    Returns:
        {
          "format": "excel",
          "file_path": str,
          "file_name": str,
          "success": bool,
          "error": str | None,
        }
    """
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except ImportError:
        return {"format": "excel", "success": False,
                "error": "openpyxl not installed.", "file_path": "", "file_name": ""}

    try:
        report_json = _build_report_json(
            query=query,
            pages=pages,
            tool_result=tool_result,
            answer_text=answer_text,
            chat_history=chat_history,
            model=model,
        )
    except Exception as exc:
        return {
            "format": "excel",
            "file_path": "",
            "file_name": "",
            "success": False,
            "error": str(exc),
        }

    _out_dir = Path(output_dir) if output_dir else REPORTS_DIR
    _out_dir.mkdir(parents=True, exist_ok=True)
    slug = re.sub(r"[^\w\-]", "_", query[:40]).strip("_")
    uniq = datetime.now().strftime("%Y%m%d_%H%M%S") + "_" + uuid.uuid4().hex[:6]
    xlsx_path = _out_dir / f"report_{slug}_{uniq}.xlsx"

    wb = Workbook()
    _hdr = Font(bold=True, color="FFFFFF")
    _hdr_fill = PatternFill("solid", fgColor="1A3A6B")
    _sub_hdr_fill = PatternFill("solid", fgColor="2D5A88")
    _alt_fill = PatternFill("solid", fgColor="F4F7FB")
    _thin = Side(style="thin", color="D0D7E2")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    # ── Keep one main data sheet, plus summary/source sheets ───────────────
    tables = report_json.get("tables", []) if isinstance(report_json, dict) else []

    if not tables:
        return {
            "format": "excel",
            "file_path": "",
            "file_name": "",
            "success": False,
            "error": "Gemini returned empty tables.",
        }

    table = tables[-1] if isinstance(tables[-1], dict) else {}
    rows = table.get("rows", []) if isinstance(table, dict) else []
    cols = table.get("columns", []) if isinstance(table, dict) else []
    if not cols and rows:
        cols = list(rows[0].keys())
    if not cols:
        cols = ["Value"]
        rows = [{"Value": "No data"}]

    used_sheet_names: set[str] = set()

    ws_summary = wb.active
    ws_summary.title = _safe_sheet_name("Summary", used_sheet_names)
    ws_summary["A1"] = report_json.get("title", f"CAD Report — {query[:80]}")
    ws_summary["A1"].font = Font(bold=True, size=14, color="1A3A6B")
    ws_summary["A3"] = "Query"
    ws_summary["B3"] = query
    ws_summary["A4"] = "Source Strategy"
    ws_summary["B4"] = str(report_json.get("source_strategy", "hybrid"))
    ws_summary["A5"] = "Decision Note"
    ws_summary["B5"] = str(report_json.get("decision_note", ""))
    ws_summary["A6"] = "Summary"
    ws_summary["B6"] = str(report_json.get("summary", ""))
    for row in range(3, 7):
        ws_summary[f"A{row}"].font = Font(bold=True, color="FFFFFF")
        ws_summary[f"A{row}"].fill = _sub_hdr_fill
        ws_summary[f"A{row}"].border = _border
        ws_summary[f"B{row}"].border = _border
        ws_summary[f"B{row}"].alignment = Alignment(vertical="top", wrap_text=True)
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 100

    ws = wb.create_sheet(_safe_sheet_name(str(table.get("sheet_name", "Data")), used_sheet_names))
    ws.append(cols)
    for cell in ws[1]:
        cell.font = _hdr
        cell.fill = _hdr_fill
        cell.border = _border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        if isinstance(row, dict):
            ws.append([row.get(c, "") for c in cols])
        else:
            ws.append(["" for _ in cols])

    for ridx in range(2, ws.max_row + 1):
        for cidx in range(1, ws.max_column + 1):
            cell = ws.cell(ridx, cidx)
            cell.border = _border
            if ridx % 2 == 0:
                cell.fill = _alt_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for i, col in enumerate(cols, start=1):
        letter = get_column_letter(i)
        max_len = len(str(col))
        for ridx in range(2, min(ws.max_row, 302) + 1):
            max_len = max(max_len, len(str(ws.cell(ridx, i).value or "")))
        ws.column_dimensions[letter].width = min(max(max_len + 3, 14), 50)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    page_rows = report_json.get("pages", []) if isinstance(report_json, dict) else []
    if isinstance(page_rows, list) and page_rows:
        ws_src = wb.create_sheet(_safe_sheet_name("Source Pages", used_sheet_names))
        src_cols = ["Page", "File", "Summary"]
        ws_src.append(src_cols)
        for cell in ws_src[1]:
            cell.font = _hdr
            cell.fill = _hdr_fill
            cell.border = _border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for entry in page_rows[:500]:
            if not isinstance(entry, dict):
                continue
            ws_src.append([
                entry.get("page_number", ""),
                entry.get("file_name", ""),
                entry.get("summary", ""),
            ])
        for ridx in range(2, ws_src.max_row + 1):
            for cidx in range(1, ws_src.max_column + 1):
                cell = ws_src.cell(ridx, cidx)
                cell.border = _border
                if ridx % 2 == 0:
                    cell.fill = _alt_fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws_src.column_dimensions["A"].width = 10
        ws_src.column_dimensions["B"].width = 44
        ws_src.column_dimensions["C"].width = 90
        ws_src.freeze_panes = "A2"
        ws_src.auto_filter.ref = ws_src.dimensions

    try:
        wb.save(str(xlsx_path))
        return {
            "format":    "excel",
            "file_path": str(xlsx_path),
            "file_name": xlsx_path.name,
            "success":   True,
            "error":     None,
        }
    except Exception as exc:
        return {
            "format":    "excel",
            "file_path": "",
            "file_name": "",
            "success":   False,
            "error":     str(exc),
        }
